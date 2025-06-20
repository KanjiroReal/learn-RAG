import base64
import json
import os
from typing import List
from io import BytesIO
from dotenv import load_dotenv

import pandas as pd 
import numpy as np
from skimage.measure import label, regionprops
from openpyxl import load_workbook
from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from docx.oxml.ns import qn
from sentence_transformers import util
from unstructured.partition.pdf import partition_pdf
from unstructured_pytesseract import pytesseract
from PIL import Image

from _agents import agent_manager, get_embedding
from _config import ModelType
from _logger import logger
from _prompts import PROMPT, prompt_manager

load_dotenv()
pytesseract.tesseract_cmd = os.getenv("TESSERACT_BACKEND_URL")

class Parser:
    def __init__(self) -> None:
        self.embedding = get_embedding()
        
    def extract_texts_from_dir(self, dir_path: str) -> List[str]:
        """extract text from all file in a directory, this method skip nested directory
        Returns:
            full_extracted_texts: all texts extracted from all file in directory as a list
        """
        
        full_extracted_texts = []
        for filename in os.listdir(dir_path):
            filepath = os.path.join(dir_path,  filename)
            if os.path.isfile(filepath):
                # docx
                if filepath.endswith(".docx"):
                    text_list = self.extract_texts_from_docx(filepath)
                # pdf
                elif filepath.endswith(".pdf"):
                    text_list =self.extract_texts_from_pdf(filepath)
                # excel
                elif filepath.endswith(".xlsx"):
                    text_list = self.extract_texts_from_excel(filepath)
                # txt
                elif filepath.endswith(".txt"):
                    text_list = self.extract_texts_from_txt(filepath)
                else:
                    logger.warning(f"Tìm thấy file không được hỗ trợ trong directory: {filepath}")
                    text_list = []
                # add to full text
                full_extracted_texts += text_list
            
        return full_extracted_texts
    
    def extract_texts_from_docx(self, file_path: str) -> List[str]:
        """
        Xử lý văn bản đầu vào là docx, call llm summarize khi gặp table, ảnh. trả về list text là các document mỗi khi user ấn <Enter> trong document.
        """
        logger.info("Đang xử lý 1 file docx...")
        doc = Document(file_path)
        file_header = f"File: {file_path} {'#'* 100}"
        # header for file
        full_text = [file_header]
        for element in doc.element.body:
            if element.tag.endswith('p'): # paragraph tag
                paragraph = Paragraph(element, doc)
                images = self._extract_images_from_paragraph(paragraph, doc)
                has_text = paragraph.text.strip()
                if has_text or images:
                    # texts
                    if has_text:
                        full_text.append(paragraph.text.strip())
                    
                    # process images
                    for image in images:
                        # image type are: bytes
                        image_ocr_data = extract_image_ocr_data(image)
                        # bytes -> base64
                        image_base64 = base64.b64encode(image).decode('utf-8')
                        image_summary = self._summary_image(image_base64=image_base64, ocr_data=image_ocr_data)
                        if image_summary:
                            full_text.append(image_summary)
            
            elif element.tag.endswith('tbl'): # table tag
                table = Table(element, doc)
                
                # I think convert the table into formated-text could be better than summary
                logger.info("Đã tìm thấy 1 bảng trong paragraph.")
                converted_table = self._convert_table_to_json(table)
                if converted_table:
                    full_text.append(converted_table)
        
        logger.success(f"Đã hoàn thành trích xuất {len(full_text)} paragraphs.")
        return full_text
    
    def extract_texts_from_excel(self, file_path: str) -> List[str]:
        """Extract excel file.
        Pipeline:
        Read -> normalize (fill merge, fomular) -> extract multiple tables -> convert json
        """
        logger.info("Đang xử lý 1 file excel...")
        df = load_merged_excel(file_path)
        tables_list = extract_table_islands(df)
        # header for f
        file_header = f"File: {file_path} {'#'* 100}"
        table_texts = [file_header] + [self._convert_table_to_json(tbl) for tbl in tables_list]
        logger.success(f"Đã hoàn thành trích xuất {len(table_texts)} bảng.")
        return table_texts
    
    def extract_texts_from_txt(self, file_path: str) -> List[str]:
        logger.info("Đang xử lý 1 file txt...")
        with open(file_path, 'r', encoding='utf-8') as f:
            data = f.read()
        file_header = f"File: {file_path} {'#'* 100}"
        texts = [file_header]
        texts += data.split('\n')
        logger.success(f"Đã hoàn thành trích xuất {len(texts)} paragraphs.")
        return texts
    
    def extract_texts_from_pdf(self, file_path: str) -> List[str]:
        """Extract list of text, table and image summary from pdf file"""
        languages = ['vie', 'eng']
        file_header = f"File: {file_path} {'#'* 100}"
        text_list = [file_header]
        logger.info("Đang Xử lý pdf...")
        elements = partition_pdf(
            filename=file_path,
            strategy='hi_res',
            extract_images_in_pdf=True,
            extract_image_block_types=["Image", "Table"],
            extract_image_block_to_payload=True,
            languages=languages
        )
        
        for element in elements:
            # text
            if getattr(element, "category", None) in ["NarrativeText", "Title", "ListItem"]:
                text_list.append(element.text)
                logger.success(f"Trích xuất thành công 1 text element.")
            
            # table 
            elif hasattr(element, "metadata"):
                if "image_base64" in element.metadata.fields:
                    image_base64 = element.metadata.fields["image_base64"]
                    logger.success("Trích xuất thành công 1 ảnh hoặc bảng")
                    # base64 -> bytes
                    image_bytes = base64.b64decode(image_base64)
                    image_ocr_data = extract_image_ocr_data(image_bytes)
                    image_summary = self._summary_image(image_base64, ocr_data=image_ocr_data)
                    if image_summary:
                        text_list.append(image_summary)
            else:
                logger.warning("Tìm thấy 1 exception element chưa được thêm vào text_lists tại bruteforce elements.")
        logger.success(f"Đã hoàn thành trích xuất {len(text_list)} pargraphs.")
        return text_list
    
    def semantic_chunk(self, text_list, threshold:float=0.3):
        """create chunk by semantic method with threshold"""
        logger.info("Đang trích xuất chunk...")
        chunks = []
        current_chunk = [text_list[0]]
        embeddings = self.embedding.encode(text_list)
        for i in range(1, len(text_list)):
            sim = util.cos_sim(embeddings[i], embeddings[i-1]).item()
            
            if sim > threshold:
                current_chunk.append(text_list[i])
            else:
                chunks.append(" ".join(current_chunk))
                current_chunk = [text_list[i]]
        
        # last chunk
        if current_chunk:
            chunks.append(" ".join(current_chunk))

        logger.info(f"Đã tạo {len(chunks)} chunks từ document.")
        return chunks
    
    def create_embedding(self, texts):
        logger.info("Đang tạo embedding...")
        embeddings = self.embedding.encode(texts)
        logger.info("Đã tạo embedding.")
        return embeddings
    

# protected methods ======================================================
    def _extract_images_from_paragraph(self, paragraph, doc) -> List[bytes]:
        images = []
        
        # search for drawing elements
        for run in paragraph.runs:
            for drawing in run.element.xpath('.//w:drawing'):
                # image
                blips = drawing.xpath('.//a:blip[@r:embed]')
                for blip in blips:
                    embed_id = blip.get(qn('r:embed'))
                    if embed_id:
                        logger.info("Đã tìm thấy 1 ảnh trong paragraph.")
                        try:
                            # get image data
                            image_part = doc.part.related_parts[embed_id]
                            images.append(image_part.blob)
                        except Exception as e:
                            logger.error(f"Lỗi khi trích xuất thông tin ảnh tại private method: _extract_images_from_paragraph: {e}")
        
        return images
    
    def _summary_image(self, image_base64: str, wrap_output: bool = True, ocr_data: dict = {}) -> str:
        """call llm to summarize image_base64 to formatted text"""
        try:
            # instruction prompt
            ocr_str = json.dumps(ocr_data, ensure_ascii=False)
            OCR_CONTEXT = prompt_manager.get_prompt(PROMPT.OCR_CONTEXT)
            OCR_CONTEXT += f"OCR data:\n{ocr_str}\n\n"
            SUMMARIZER_PROMPT = prompt_manager.get_prompt(PROMPT.PARSER_SUMMARIZER)
            instruction = OCR_CONTEXT + SUMMARIZER_PROMPT
            
            # agent
            image_summarizer_agent = agent_manager.create_agent(
            name="Summarizer",
            instruction=instruction,
            model_type=ModelType.VL,
            )
            
            # input message
            image_data_url = f"data:image/jpeg;base64,{image_base64}"
            message = [
                {
                    "role": "user",
                    "content": [
                        {"type": "input_image", "image_url": image_data_url},
                        {"type": "input_text", "text": "Convert this image url to text with your instruction."}
                    ]
                }
            ]
            # run agent
            response = agent_manager.run_agent(
                agent=image_summarizer_agent, 
                prompt=message,
            )
            logger.success("Đã chuyển đổi 1 bức ảnh thành nội dung tóm tắt.")
            if wrap_output:
                return_text = f"[Đây là một bức ảnh, bức ảnh đã được thay thế bằng mô tả của AI][MÔ TẢ HÌNH ẢNH] \n{response.final_output} \n[KẾT THÚC MÔ TẢ HÌNH ẢNH]"
            else:
                return_text = response.final_output
            return return_text
        except Exception as e:
            logger.error(f"Lỗi khi call llm về hình ảnh tại method _summary_image: {e}")
            raise e.with_traceback(e.__traceback__)
        
    def _convert_table_to_markdown(self, table: pd.DataFrame | Table) -> str:
        """Convert docx.Table or pandas.DataFrame to markdown string.
        For DataFrame:
        | Col1 | Col2 |
        | ---- | ---- |
        | val1 | val2 |
        For docx.Table: first row treated as header.
        """
        # Prepare rows as lists of strings
        rows: list[list[str]] = []
        if isinstance(table, pd.DataFrame):
            df = table.reset_index(drop=True)
            # Header row
            columns = list(df.columns)
            rows.append(columns)
            # Data rows
            for _, row in df.iterrows():
                rows.append([str(row[col]) for col in columns])
        elif isinstance(table, Table):
            for row_index, row in enumerate(table.rows):
                cells = [cell.text.strip().replace('\n', ' ') for cell in row.cells]
                # For header (first row), record
                rows.append(cells)
        else:
            raise TypeError(f"Input must be either pandas.DataFrame or docx.table.Table, got {type(table)}.")
        # Build markdown
        markdown_lines = []
        if rows:
            # Header
            header_cells = rows[0]
            header_line = "| " + " | ".join(header_cells) + " |"
            # Separator
            sep_line = "|" + "|".join([" --- " for _ in header_cells]) + "|"
            markdown_lines.append(header_line)
            markdown_lines.append(sep_line)
            # Data rows
            for data_cells in rows[1:]:
                if len(data_cells) < len(header_cells):
                    data_cells += [""] * (len(header_cells) - len(data_cells))
                
                data_cells = data_cells[:len(header_cells)]
                line = "| " + " | ".join(data_cells) + " |"
                markdown_lines.append(line)
        converted = "\n".join(markdown_lines)
        logger.success("Đã chuyển đổi bảng thành markdown format.")
        return f"[ĐÂY LÀ BẢNG ĐÃ ĐƯỢC CHUYỂN ĐỔI THÀNH MARKDOWN FORMAT] \n{converted} \n[KẾT THÚC BẢNG]"

    def _convert_table_to_json(self, table: pd.DataFrame | Table) -> str:
        """Convert docx.Table or pd.Dataframe table to json
        if the table have the format
        |      | Col2 | col3 |
        | row2 | a    | c    |
        | row3 | b    | d    | 
        the json will be
        [
            {
                "key": "",
                "value": [row2, row3]
            },
            {
                "key": "Col2",
                "value": [a, b]
            },
            {
                "key": "Col3",
                "value": [c, d]
            }
        ]
        """
        if isinstance(table, pd.DataFrame):
            table = table.reset_index(drop=True)
            columns = list(table.columns)
            values = [table[col].astype(str).tolist() for col in table.columns]
        elif isinstance(table, Table):
            header = []
            values = []
            for row_index, row in enumerate(table.rows):
                if row_index == 0:  # header
                    header = [cell.text for cell in row.cells]
                    values = [[] for _ in header]
                else:
                    for col_index, cell in enumerate(row.cells):
                        if col_index < len(values):
                            values[col_index].append(cell.text)
            columns = header
        else:
            raise TypeError(f"Input must be either docx.table.Table or pandas.DataFrame, Got {type(table)}.")
        
        table_json = [{"key": key, "value": value} for key, value in zip(columns, values)]
        converted_table = json.dumps(table_json, indent=2, ensure_ascii=False)
        logger.success("Đã chuyển đổi 1 bảng thành json format.")
        return f"[ĐÂY LÀ BẢNG ĐÃ ĐƯỢC CHUYỂN ĐỔI THÀNH JSON FORMAT] \n{converted_table} \n[KẾT THÚC BẢNG]"
    
# functions ======================================================
def extract_table_islands(df: pd.DataFrame) -> List[pd.DataFrame]:
    """Extract multiple discrete table from an excel file"""
    boolean_matrix = label(np.array(df.notnull().astype("int")))
    
    list_dfs = []
    for s in regionprops(boolean_matrix):
        sub_df = (df.iloc[s.bbox[0]:s.bbox[2], s.bbox[1]:s.bbox[3]]
                    .pipe(lambda df_: df_.rename(columns=df_.iloc[0]) # type: ignore
                    .drop(df_.index[0])))
        list_dfs.append(sub_df)
    
    return list_dfs

def load_merged_excel(filepath, sheet_name=0) -> pd.DataFrame:
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name if isinstance(sheet_name, str) else wb.sheetnames[sheet_name]]
    
    # fill merged
    for merged_cell_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged_cell_range.bounds
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merged_cell_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col +1):
                if ws.cell(row=row, column=col).value is None:
                    ws.cell(row=row, column=col, value=top_left_value)
    data = ws.values
    df = pd.DataFrame(data)
    return df

def extract_image_ocr_data(image_bytes: bytes) -> dict:
    """
    Extracts all text content and corresponding bounding box positions from an image provided as a base64-encoded data URL.

    This function performs OCR (Optical Character Recognition) on the input image and returns both:
    - The full extracted text as a single string.
    - A list of individual text blocks, each with its corresponding bounding box coordinates.

    Args:
        image (bytes): image in bytes

    Returns:
        dict: A dictionary with the following keys:
            - 'text' (str): The complete extracted text, concatenated from all detected text blocks.
            - 'text_blocks' (List[dict]): A list of dictionaries, each representing a text block with:
                - 'text' (str): The recognized text content of the block.
                - 'left' (int): The x-coordinate of the top-left corner of the bounding box.
                - 'top' (int): The y-coordinate of the top-left corner of the bounding box.
                - 'width' (int): The width of the bounding box.
                - 'height' (int): The height of the bounding box.
    """
    image = Image.open(BytesIO(image_bytes))
    
    languages = ['vie', 'eng', 'enm']
    ocr_data = pytesseract.image_to_data(image, lang="+".join(languages), output_type=pytesseract.Output.DICT)
    blocks = []
    for i in range(len(ocr_data['text'])):
        text = ocr_data['text'][i].strip()
        if text:
            block = {
                'text': text,
                'left': ocr_data['left'][i],
                'top': ocr_data['top'][i],
                'width': ocr_data['width'][i],
                'height': ocr_data['height'][i]
            }
            blocks.append(block)
    
    full_text = ' '.join([block['text'] for block in blocks])
    
    output = {
        'text': full_text,
        'text_blocks': blocks
    }
    logger.debug(f"OCR output: \n\n{output}")
    
    return output